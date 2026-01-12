from __future__ import annotations

import csv
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

from .phones import format_phone_number


_COLUMN_MAP: dict[str, str] = {
    "name": "name",
    "الاسم": "name",
    "full name": "name",
    "الاسم الكامل": "name",
    "phone": "phone",
    "الهاتف": "phone",
    "رقم الهاتف": "phone",
    "mobile": "phone",
    "الجوال": "phone",
    "email": "email",
    "البريد الإلكتروني": "email",
    "الايميل": "email",
    "title": "title",
    "المسمى الوظيفي": "title",
    "job title": "title",
    "الوظيفة": "title",
    "company": "company",
    "الشركة": "company",
    "organization": "company",
    "المؤسسة": "company",
    "industry": "industry",
    "المجال": "industry",
    "sector": "industry",
    "professional_bio": "professional_bio",
    "نبذة مهنية": "professional_bio",
    "bio": "professional_bio",
    "personal_bio": "personal_bio",
    "نبذة شخصية": "personal_bio",
    "skills": "skills",
    "المهارات": "skills",
    "expertise": "skills",
    "looking_for": "looking_for",
    "يبحث عن": "looking_for",
    "looking for": "looking_for",
    "offering": "offering",
    "يقدم": "offering",
    "can offer": "offering",
    "linkedin": "linkedin_url",
    "linkedin_url": "linkedin_url",
    "لينكدإن": "linkedin_url",
    "photo": "photo_url",
    "photo_url": "photo_url",
    "الصورة": "photo_url",
    "location": "location",
    "الموقع": "location",
    "city": "location",
    "languages": "languages",
    "اللغات": "languages",
}


def _normalize_col(col: str) -> str:
    return " ".join(col.lower().strip().replace("_", " ").replace("-", " ").split())


def _map_columns(columns: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for col in columns:
        norm = _normalize_col(col)
        if norm in _COLUMN_MAP:
            mapping[col] = _COLUMN_MAP[norm]
    return mapping


def _parse_rows(rows: list[dict], event_id: str) -> list[dict]:
    if not rows:
        raise ValueError("الملف فارغ")

    columns = list(rows[0].keys())
    col_map = _map_columns(columns)
    mapped_fields = set(col_map.values())
    if "name" not in mapped_fields:
        raise ValueError('الملف يجب أن يحتوي على عمود "الاسم"')
    if "phone" not in mapped_fields:
        raise ValueError('الملف يجب أن يحتوي على عمود "الهاتف"')

    attendees: list[dict] = []
    seen_phones: set[str] = set()

    for row in rows:
        attendee: dict = {"id": str(uuid4()), "event_id": event_id}
        for col, field in col_map.items():
            value = row.get(col)
            if value is None or value == "":
                continue
            v = str(value).strip()
            if not v:
                continue
            if field == "phone":
                v = format_phone_number(v)
            attendee[field] = v

        if not attendee.get("name") or not attendee.get("phone"):
            continue
        if attendee["phone"] in seen_phones:
            continue
        seen_phones.add(attendee["phone"])
        attendees.append(attendee)

    return attendees


def parse_attendees_file(file_path: Path, event_id: str) -> list[dict]:
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        with file_path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        return _parse_rows(rows, event_id)

    # Default: Excel
    wb = load_workbook(filename=str(file_path), read_only=True)
    ws = wb[wb.sheetnames[0]]
    values = list(ws.values)
    if not values:
        raise ValueError("الملف فارغ")
    headers = [str(h).strip() if h is not None else "" for h in values[0]]
    rows: list[dict] = []
    for r in values[1:]:
        row = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row[header] = r[idx] if idx < len(r) else None
        rows.append(row)
    return _parse_rows(rows, event_id)


